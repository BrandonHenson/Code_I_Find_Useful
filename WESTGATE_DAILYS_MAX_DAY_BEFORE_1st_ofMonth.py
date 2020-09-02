import openpyxl
import os
import keyboard
import time

filelist = open(str(os.getcwd()) + '/MAXDAY.xls', 'a')
input('Run this from the folder with the daily files in it')
# remove irrelevant sheets
print("Deleting All Sheets Except Westgate")
ind = 1
for i in os.listdir(os.getcwd()):
    if 'xlsx' in i:
        print("Deleting from file " + str(ind))
        wb = openpyxl.load_workbook(i)
        del wb['System Flow Summary']
        del wb['System Reservoir Summary']
        del wb['System Pressure Summary']
        del wb['MWWD Admin HQ']
        del wb['Mukilteo Flowmeter & FCV']
        del wb['100th St. FCV']
        del wb['Casino Road Flowmeter']
        del wb['112th St. FCV']
        del wb['Harbor Point FCV']
        del wb['Reservoir #1 & PS']
        del wb['620 Zone 2MG Reservoir (#2)']
        del wb['Paine Field Reservoir (#4)']
        del wb['620 Zone 6MG Reservoir (#5)']
        del wb['ChangeLog']
        del wb['Information']
        wb.save(i)
        ind = ind + 1
# save values of westgate sheet
print("Getting Values From Formulas")
ind = 1
for i in os.listdir(os.getcwd()):
    if 'xlsx' in i:
        print("Getting Values from file " + str(ind))
        p = os.path.join(os.getcwd() + '\\' + i)
        os.popen(p)
        time.sleep(5)
        keyboard.press_and_release('alt')
        time.sleep(.25)
        keyboard.press_and_release('l')
        time.sleep(.25)
        keyboard.press_and_release('v')
        time.sleep(.25)
        keyboard.press_and_release('f7')
        time.sleep(.25)
        keyboard.write('Sub Saveasvalue()')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.write('Dim wsh As Worksheet')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.write('For Each wsh In ThisWorkbook.Worksheets')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.write(' wsh.Cells.Copy')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.write('wsh.Cells.PasteSpecial xlPasteValues')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.write('Next')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.write('Application.CutCopyMode = False')
        time.sleep(.25)
        keyboard.write('\n')
        time.sleep(.25)
        keyboard.press_and_release('f5')
        time.sleep(15)
        keyboard.press_and_release('alt + q')
        time.sleep(2)
        keyboard.press_and_release('alt')
        time.sleep(.25)
        keyboard.press_and_release('f')
        time.sleep(.25)
        keyboard.press_and_release('s')
        time.sleep(.25)
        keyboard.press_and_release('enter')
        time.sleep(.25)
        keyboard.press_and_release('enter')
        time.sleep(2)
        os.system("taskkill /f /im EXCEL.EXE")
        ind = ind + 1
print("Displaying Date And Values (Last Step)")
# Display file and day average
ind = 1
for i in os.listdir(os.getcwd()):
    if 'xlsx' in i:
        print("Writing Values from  file " + str(ind))
        My_path = os.getcwd() + '\\' + i
        wb_obj = openpyxl.load_workbook(My_path, data_only=True)
        my_sheet_obj = wb_obj.active
        my_cell_obj = my_sheet_obj.cell(row=30, column=16)
        i = i.replace('.xlsx', '')
        # print(i.replace('DailyWaterReport_','') + ' '+ str(my_cell_obj.value))
        filelist.writelines(str(i.replace('DailyWaterReport_', '') + '\t' + str(my_cell_obj.value) + '\n'))
        ind = ind + 1
