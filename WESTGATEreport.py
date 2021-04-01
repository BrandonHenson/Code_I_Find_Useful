import openpyxl
import os
import keyboard
import time




for i in os.listdir(os.getcwd()):
    os.rename(i,i.replace(' ','_'))
for i in os.listdir(os.getcwd()):
    os.rename(i, i.replace('(', ''))
for i in os.listdir(os.getcwd()):
    os.rename(i, i.replace(')', ''))
for i in os.listdir(os.getcwd()):
    if '.pdf' in i:
        os.popen(i)
        time.sleep(3)
        keyboard.press_and_release('ctrl + shift + s')
        time.sleep(.25)
        keyboard.write('PF_CONS_DATA')
        time.sleep(.25)
        keyboard.press_and_release('tab')
        time.sleep(.25)
        keyboard.press_and_release('down')
        time.sleep(.25)
        keyboard.press_and_release('down')
        time.sleep(.25)
        keyboard.press_and_release('down')
        time.sleep(.25)
        keyboard.press_and_release('enter')
        time.sleep(.25)
        keyboard.press_and_release('enter')
        time.sleep(2)
        os.system("taskkill /f /im Acrobat.exe")
        p = i.replace('.pdf','.xlsx')
        time.sleep(2)
list = []
My_path2 = os.getcwd() + '\\' + 'Discharge_Report_AutoTemplate.xlsx'
My_path = os.getcwd() + '\\' + 'PF_CONS_DATA.xlsx'
wb_obj = openpyxl.load_workbook(My_path, data_only=True)
my_sheet_obj = wb_obj.active
subtotal = my_sheet_obj.cell(row=12, column=12).value
effmeter = my_sheet_obj.cell(row=13, column=12).value
date = my_sheet_obj.cell(row=2, column=3).value
wb_obj2 = openpyxl.load_workbook(My_path2, data_only=True)
my_sheet_obj2 = wb_obj2.active
my_sheet_obj2['B12'] = effmeter
my_sheet_obj2['B13'] = subtotal
my_sheet_obj2['B8'] = date
wb_obj2.save(filename = 'AutoTemplateFILLEDIN.xlsx')
